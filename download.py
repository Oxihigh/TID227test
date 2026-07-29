import os
import time
import glob
import logging
import shutil
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

NAVY_FILL = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
LIGHT_NAVY_FILL = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
SILVER_FILL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")

FONT_HEADER_WHITE = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Segoe UI", size=10, color="2D3748")
FONT_BODY_BOLD = Font(name="Segoe UI", size=10, bold=True, color="1A202C")
FONT_SUBTITLE = Font(name="Segoe UI", size=12, italic=True, color="718096")

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

THIN_BORDER_SIDE = Side(border_style="thin", color="CBD5E0")
THIN_BORDER = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)

def setup_logging():
    for log_file in glob.glob("*.log"):
        try:
            os.remove(log_file)
        except Exception:
            pass
            
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    log_filename = f"run_{timestamp}.log"
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filename, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    return log_filename

TARGET_PAGES = {
    "district_performance": "https://dashboards.toastmasters.org/District.aspx?id=227",
    "division_area_performance": "https://dashboards.toastmasters.org/Division.aspx?id=227",
    "club_performance": "https://dashboards.toastmasters.org/Club.aspx?id=227"
}

def download_data_from_url(page, name, url, timestamp):
    logging.info(f"🌐 Navigating to {name}...")
    page.goto(url, timeout=60000)
    page.wait_for_load_state("networkidle")
    time.sleep(2)
    
    logging.info(f"📥 Triggering CSV download for {name}...")
    with page.expect_download() as download_info:
        page.select_option("select[name*='Export']", value="CSV")
        
    download = download_info.value
    data_dir = os.path.join(os.getcwd(), "data")
    os.makedirs(data_dir, exist_ok=True)
    
    save_path = os.path.join(data_dir, f"{name}_{timestamp}.csv")
    download.save_as(save_path)
    logging.info(f"✅ Download completed: {save_path}")
    return save_path

def style_range(sheet, cell_range, font=None, alignment=None, fill=None, border=None):
    for row in sheet[cell_range]:
        for cell in row:
            if font: cell.font = font
            if alignment: cell.alignment = alignment
            if fill: cell.fill = fill
            if border: cell.border = border

def autofit_columns(sheet):
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if val.startswith('='):
                val_len = 12
            else:
                val_len = len(val)
            max_len = max(max_len, val_len)
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 10)

def load_and_clean_data(data_dir, timestamp):
    club_files = glob.glob(os.path.join(data_dir, f"club_performance_{timestamp}.csv"))
    district_files = glob.glob(os.path.join(data_dir, f"district_performance_{timestamp}.csv"))
    div_files = glob.glob(os.path.join(data_dir, f"division_area_performance_{timestamp}.csv"))

    if not (club_files and district_files and div_files):
        logging.error("❌ Missing downloaded performance CSV files.")
        return None, []

    df_cp = pd.read_csv(club_files[0]).dropna(subset=['Club Name'])
    df_dp = pd.read_csv(district_files[0]).dropna(subset=['Club Name'])
    df_div = pd.read_csv(div_files[0]).dropna(subset=['Club Name'])

    df_cp['Club Number'] = df_cp['Club Number'].astype(int)
    df_dp['Club'] = df_dp['Club'].astype(int)
    df_div['Club'] = df_div['Club'].astype(int)

    df_dp = df_dp.rename(columns={'Club': 'Club Number'})
    df_div = df_div.rename(columns={'Club': 'Club Number'})

    for df in [df_cp, df_dp, df_div]:
        df['Division'] = df['Division'].astype(str).str.strip()
        df['Area'] = df['Area'].astype(str).str.strip()

    keys = ['District', 'Division', 'Area', 'Club Number', 'Club Name']
    merged = df_cp.merge(df_dp, on=keys, how='outer', suffixes=('', '_dp'))
    merged = merged.merge(df_div, on=keys, how='outer', suffixes=('', '_div'))

    cols = [
        'Club Number', 'District', 'Division', 'Area', 'Club Name', 'Club Status',
        'Mem. Base', 'Active Members', 'Goals Met',
        'New', 'Oct. Ren.', 'Apr. Ren.', 'Total to Date'
    ]
    
    for col in cols:
        if col not in merged.columns:
            merged[col] = 0
            
    final_df = merged[cols].copy()
    final_df = final_df.rename(columns={
        'Mem. Base': 'Base Membership',
        'Active Members': 'Active Membership',
        'New': 'New Member Payments',
        'Oct. Ren.': 'September Renewals',
        'Apr. Ren.': 'March Renewals',
        'Total to Date': 'Total Payments'
    })
    
    # Calculate Single vs Double Renewals logic for September & March Renewal Report
    final_df['Sept Single Renewal'] = final_df['September Renewals'].apply(lambda x: max(0, int(x)))
    final_df['Sept Double Renewal'] = 0
    final_df['March Single Renewal'] = final_df['March Renewals'].apply(lambda x: max(0, int(x)))
    final_df['March Double Renewal'] = 0
    
    final_df = final_df.sort_values(by=['Division', 'Area', 'Club Number']).reset_index(drop=True)
    return final_df, club_files + district_files + div_files

def generate_excel_mastersheet(final_df, output_excel_path):
    logging.info("[INFO] Generating Multi-Sheet Excel Mastersheet...")
    wb = openpyxl.Workbook()
    
    # 1. Setup ClubDetails sheet
    ws_details = wb.active
    ws_details.title = "ClubDetails"
    ws_details.views.sheetView[0].showGridLines = True
    
    headers = list(final_df.columns)
    ws_details.append(headers)
    style_range(ws_details, f"A1:{get_column_letter(len(headers))}1", font=FONT_HEADER_WHITE, alignment=ALIGN_CENTER, fill=NAVY_FILL, border=THIN_BORDER)
    
    for idx, row in final_df.iterrows():
        ws_details.append(list(row.values))
        row_num = idx + 2
        fill = ZEBRA_FILL if row_num % 2 == 0 else WHITE_FILL
        style_range(ws_details, f"A{row_num}:{get_column_letter(len(headers))}{row_num}", font=FONT_BODY, alignment=ALIGN_LEFT, fill=fill, border=THIN_BORDER)
        ws_details[f"A{row_num}"].alignment = ALIGN_CENTER
        ws_details[f"B{row_num}"].alignment = ALIGN_CENTER
        ws_details[f"C{row_num}"].alignment = ALIGN_CENTER
        ws_details[f"D{row_num}"].alignment = ALIGN_CENTER
        ws_details[f"F{row_num}"].alignment = ALIGN_CENTER
        for c in ["G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q"]:
            ws_details[f"{c}{row_num}"].number_format = "#,##0"
            ws_details[f"{c}{row_num}"].alignment = ALIGN_RIGHT
            
    ws_details.column_dimensions['A'].width = 15
    ws_details.column_dimensions['E'].width = 30
    autofit_columns(ws_details)

    divisions = sorted(final_df['Division'].dropna().unique().tolist())
    
    # 2. Division sheets
    for div in divisions:
        sheet_name = f"Division {div}"
        ws_div = wb.create_sheet(title=sheet_name)
        ws_div.views.sheetView[0].showGridLines = True
        
        ws_div.cell(row=1, column=1, value=f"Division {div} - Membership and Payments Performance").font = Font(name="Segoe UI", size=14, bold=True, color="1B365D")
        
        headers_div = [
            "Club Number", "Area", "Club Name", "Base Membership", "Active Membership", 
            "DCP Goals Met", "Distinguished Status", 
            "September Renewals", "September Renewals %", 
            "March Renewals", "March Renewals %", 
            "Total Payments", "Total New Members",
            "July", "August", "September", "October", "November", "December", 
            "January", "February", "March", "April", "May", "June",
            "Smedley Award Eligibility", "Smedley Award Goal",
            "Talk Up Eligibility", "Talk Up Goal",
            "Beat the Clock Eligibility", "Beat the Clock Goal"
        ]
        
        for col_idx, header in enumerate(headers_div, 1):
            cell = ws_div.cell(row=3, column=col_idx, value=header)
            cell.font = FONT_HEADER_WHITE
            cell.fill = NAVY_FILL
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
            
        div_clubs = final_df[final_df['Division'] == div].copy()
        
        for r_idx, (df_idx, row) in enumerate(div_clubs.iterrows(), 4):
            ws_div.cell(row=r_idx, column=1, value=row['Club Number']).alignment = ALIGN_CENTER
            ws_div.cell(row=r_idx, column=2, value=row['Area']).alignment = ALIGN_CENTER
            ws_div.cell(row=r_idx, column=3, value=row['Club Name']).alignment = ALIGN_LEFT
            
            club_row_in_details = df_idx + 2
            
            ws_div.cell(row=r_idx, column=4, value=f"=ClubDetails!G{club_row_in_details}")
            ws_div.cell(row=r_idx, column=5, value=f"=ClubDetails!H{club_row_in_details}")
            ws_div.cell(row=r_idx, column=6, value=f"=ClubDetails!I{club_row_in_details}")
            ws_div.cell(row=r_idx, column=7, value=f'=IF(F{r_idx}<5,"-",IF(E{r_idx}>19,"Yes",IF(E{r_idx}>D{r_idx}+2,"Yes","-")))')
            ws_div.cell(row=r_idx, column=8, value=f"=ClubDetails!K{club_row_in_details}")
            ws_div.cell(row=r_idx, column=9, value=f"=IF(D{r_idx}=0,0,H{r_idx}/D{r_idx})")
            ws_div.cell(row=r_idx, column=10, value=f"=ClubDetails!L{club_row_in_details}")
            ws_div.cell(row=r_idx, column=11, value=f"=IF(D{r_idx}=0,0,J{r_idx}/D{r_idx})")
            ws_div.cell(row=r_idx, column=12, value=f"=ClubDetails!M{club_row_in_details}")
            ws_div.cell(row=r_idx, column=13, value=f"=ClubDetails!J{club_row_in_details}")
            
            for col_idx in range(14, 26):
                ws_div.cell(row=r_idx, column=col_idx, value=0).number_format = "#,##0"
                ws_div.cell(row=r_idx, column=col_idx).alignment = ALIGN_RIGHT
                
            ws_div.cell(row=r_idx, column=26, value=f'=IF(SUM(O{r_idx}:P{r_idx})>=5,"Yes","No")')
            ws_div.cell(row=r_idx, column=27, value=f'=IF(SUM(O{r_idx}:P{r_idx})>=5,0,5-SUM(O{r_idx}:P{r_idx}))')
            ws_div.cell(row=r_idx, column=28, value=f'=IF(SUM(U{r_idx}:V{r_idx})>=5,"Yes","No")')
            ws_div.cell(row=r_idx, column=29, value=f'=IF(SUM(U{r_idx}:V{r_idx})>=5,0,5-SUM(U{r_idx}:V{r_idx}))')
            ws_div.cell(row=r_idx, column=30, value=f'=IF(SUM(X{r_idx}:Y{r_idx})>=5,"Yes","No")')
            ws_div.cell(row=r_idx, column=31, value=f'=IF(SUM(X{r_idx}:Y{r_idx})>=5,0,5-SUM(X{r_idx}:Y{r_idx}))')
            
            fill = ZEBRA_FILL if r_idx % 2 == 0 else WHITE_FILL
            for c_idx in range(1, 32):
                cell = ws_div.cell(row=r_idx, column=c_idx)
                cell.font = FONT_BODY
                cell.border = THIN_BORDER
                if c_idx not in [1, 2, 3, 7, 26, 28, 30]:
                    cell.alignment = ALIGN_RIGHT
                    if c_idx in [9, 11]:
                        cell.number_format = "0.0%"
                    else:
                        cell.number_format = "#,##0"
                elif c_idx in [7, 26, 28, 30]:
                    cell.alignment = ALIGN_CENTER

        summary_row = len(div_clubs) + 4
        ws_div.cell(row=summary_row, column=1, value="Total / Average").font = FONT_BODY_BOLD
        ws_div.cell(row=summary_row, column=1).alignment = ALIGN_LEFT
        ws_div.cell(row=summary_row, column=1).border = THIN_BORDER
        ws_div.cell(row=summary_row, column=1).fill = SILVER_FILL
        
        ws_div.cell(row=summary_row, column=2, value="").border = THIN_BORDER
        ws_div.cell(row=summary_row, column=2).fill = SILVER_FILL
        ws_div.cell(row=summary_row, column=3, value="").border = THIN_BORDER
        ws_div.cell(row=summary_row, column=3).fill = SILVER_FILL
        
        ws_div.cell(row=summary_row, column=4, value=f"=SUM(D4:D{summary_row-1})")
        ws_div.cell(row=summary_row, column=5, value=f"=SUM(E4:E{summary_row-1})")
        ws_div.cell(row=summary_row, column=6, value=f"=AVERAGE(F4:F{summary_row-1})")
        ws_div.cell(row=summary_row, column=7, value=f'=COUNTIF(G4:G{summary_row-1},"Yes")')
        ws_div.cell(row=summary_row, column=8, value=f"=SUM(H4:H{summary_row-1})")
        ws_div.cell(row=summary_row, column=9, value=f"=IF(D{summary_row}=0,0,H{summary_row}/D{summary_row})")
        ws_div.cell(row=summary_row, column=10, value=f"=SUM(J4:J{summary_row-1})")
        ws_div.cell(row=summary_row, column=11, value=f"=IF(D{summary_row}=0,0,J{summary_row}/D{summary_row})")
        ws_div.cell(row=summary_row, column=12, value=f"=SUM(L4:L{summary_row-1})")
        ws_div.cell(row=summary_row, column=13, value=f"=SUM(M4:M{summary_row-1})")
        
        for m_col in range(14, 26):
            m_letter = get_column_letter(m_col)
            ws_div.cell(row=summary_row, column=m_col, value=f"=SUM({m_letter}4:{m_letter}{summary_row-1})")
            
        ws_div.cell(row=summary_row, column=26, value=f'=COUNTIF(Z4:Z{summary_row-1},"Yes")')
        ws_div.cell(row=summary_row, column=27, value=f"=SUM(AA4:AA{summary_row-1})")
        ws_div.cell(row=summary_row, column=28, value=f'=COUNTIF(AB4:AB{summary_row-1},"Yes")')
        ws_div.cell(row=summary_row, column=29, value=f"=SUM(AC4:AC{summary_row-1})")
        ws_div.cell(row=summary_row, column=30, value=f'=COUNTIF(AD4:AD{summary_row-1},"Yes")')
        ws_div.cell(row=summary_row, column=31, value=f"=SUM(AE4:AE{summary_row-1})")
        
        for c_idx in range(4, 32):
            cell = ws_div.cell(row=summary_row, column=c_idx)
            cell.font = FONT_BODY_BOLD
            cell.border = THIN_BORDER
            cell.fill = SILVER_FILL
            if c_idx not in [7, 26, 28, 30]:
                cell.alignment = ALIGN_RIGHT
                if c_idx in [9, 11]:
                    cell.number_format = "0.0%"
                elif c_idx == 6:
                    cell.number_format = "0.0"
                else:
                    cell.number_format = "#,##0"
            else:
                cell.alignment = ALIGN_CENTER
                
        ws_div.column_dimensions['A'].width = 15
        ws_div.column_dimensions['C'].width = 25
        autofit_columns(ws_div)

    # 3. Overall Sheet
    ws_ov = wb.create_sheet(title="Overall", index=1)
    ws_ov.views.sheetView[0].showGridLines = True
    ws_ov.cell(row=1, column=1, value="District 227 - Divisions Performance Rolled-up Summary").font = Font(name="Segoe UI", size=14, bold=True, color="1B365D")
    
    headers_ov_div = [
        "Division", "Base Clubs", "Active Clubs", "Distinguished Clubs", 
        "Base Membership", "Active Membership", 
        "September Renewals", "September Renewals %", 
        "March Renewals", "March Renewals %", 
        "Total New Members", "Total Payments", "20+ Clubs"
    ]
    
    for c_idx, header in enumerate(headers_ov_div, 1):
        cell = ws_ov.cell(row=3, column=c_idx, value=header)
        cell.font = FONT_HEADER_WHITE
        cell.fill = NAVY_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        
    for r_idx, div in enumerate(divisions, 4):
        ws_ov.cell(row=r_idx, column=1, value=f"Division {div}").alignment = ALIGN_LEFT
        div_sheet_ref = f"'Division {div}'"
        num_clubs_in_div = len(final_df[final_df['Division'] == div])
        div_last_row = num_clubs_in_div + 3
        
        ws_ov.cell(row=r_idx, column=2, value=f"=COUNTIFS(ClubDetails!C:C, \"{div}\")")
        ws_ov.cell(row=r_idx, column=3, value=f"=COUNTIFS(ClubDetails!C:C, \"{div}\", ClubDetails!F:F, \"Active\")")
        ws_ov.cell(row=r_idx, column=4, value=f"={div_sheet_ref}!G{div_last_row+1}")
        ws_ov.cell(row=r_idx, column=5, value=f"={div_sheet_ref}!D{div_last_row+1}")
        ws_ov.cell(row=r_idx, column=6, value=f"={div_sheet_ref}!E{div_last_row+1}")
        ws_ov.cell(row=r_idx, column=7, value=f"={div_sheet_ref}!H{div_last_row+1}")
        ws_ov.cell(row=r_idx, column=8, value=f"=IF(E{r_idx}=0, 0, G{r_idx}/E{r_idx})")
        ws_ov.cell(row=r_idx, column=9, value=f"={div_sheet_ref}!K{div_last_row+1}")
        ws_ov.cell(row=r_idx, column=10, value=f"=IF(E{r_idx}=0, 0, I{r_idx}/E{r_idx})")
        ws_ov.cell(row=r_idx, column=11, value=f"={div_sheet_ref}!O{div_last_row+1}")
        ws_ov.cell(row=r_idx, column=12, value=f"={div_sheet_ref}!N{div_last_row+1}")
        ws_ov.cell(row=r_idx, column=13, value=f"=COUNTIF({div_sheet_ref}!E4:E{div_last_row}, \">=20\")")
        
        fill = ZEBRA_FILL if r_idx % 2 == 0 else WHITE_FILL
        for c_idx in range(1, 14):
            cell = ws_ov.cell(row=r_idx, column=c_idx)
            cell.font = FONT_BODY
            cell.border = THIN_BORDER
            cell.fill = fill
            if c_idx > 1:
                cell.alignment = ALIGN_RIGHT
                if c_idx in [8, 10]:
                    cell.number_format = "0.0%"
                else:
                    cell.number_format = "#,##0"
                    
    tot_row = len(divisions) + 4
    ws_ov.cell(row=tot_row, column=1, value="District Total").font = FONT_BODY_BOLD
    ws_ov.cell(row=tot_row, column=1).alignment = ALIGN_LEFT
    ws_ov.cell(row=tot_row, column=1).border = THIN_BORDER
    ws_ov.cell(row=tot_row, column=1).fill = SILVER_FILL
    
    ws_ov.cell(row=tot_row, column=2, value=f"=SUM(B4:B{tot_row-1})")
    ws_ov.cell(row=tot_row, column=3, value=f"=SUM(C4:C{tot_row-1})")
    ws_ov.cell(row=tot_row, column=4, value=f"=SUM(D4:D{tot_row-1})")
    ws_ov.cell(row=tot_row, column=5, value=f"=SUM(E4:E{tot_row-1})")
    ws_ov.cell(row=tot_row, column=6, value=f"=SUM(F4:F{tot_row-1})")
    ws_ov.cell(row=tot_row, column=7, value=f"=SUM(G4:G{tot_row-1})")
    ws_ov.cell(row=tot_row, column=8, value=f"=IF(E{tot_row}=0, 0, G{tot_row}/E{tot_row})")
    ws_ov.cell(row=tot_row, column=9, value=f"=SUM(I4:I{tot_row-1})")
    ws_ov.cell(row=tot_row, column=10, value=f"=IF(E{tot_row}=0, 0, I{tot_row}/E{tot_row})")
    ws_ov.cell(row=tot_row, column=11, value=f"=SUM(K4:K{tot_row-1})")
    ws_ov.cell(row=tot_row, column=12, value=f"=SUM(L4:L{tot_row-1})")
    ws_ov.cell(row=tot_row, column=13, value=f"=SUM(M4:M{tot_row-1})")
    
    for c_idx in range(2, 14):
        cell = ws_ov.cell(row=tot_row, column=c_idx)
        cell.font = FONT_BODY_BOLD
        cell.border = THIN_BORDER
        cell.fill = SILVER_FILL
        cell.alignment = ALIGN_RIGHT
        if c_idx in [8, 10]:
            cell.number_format = "0.0%"
        else:
            cell.number_format = "#,##0"

    area_start_row = tot_row + 4
    ws_ov.cell(row=area_start_row-1, column=1, value="Area Performance Summary").font = Font(name="Segoe UI", size=12, bold=True, color="1B365D")
    
    headers_area = [
        "Division", "Area", "Base Membership", "Active Membership", 
        "September Renewals", "September Renewals %", 
        "March Renewals", "March Renewals %", 
        "Total New Members", "Total Payments"
    ]
    
    for c_idx, header in enumerate(headers_area, 1):
        cell = ws_ov.cell(row=area_start_row, column=c_idx, value=header)
        cell.font = FONT_HEADER_WHITE
        cell.fill = LIGHT_NAVY_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        
    area_combinations = final_df.groupby(['Division', 'Area']).size().reset_index()[['Division', 'Area']]
    
    for a_idx, row in area_combinations.iterrows():
        curr_row = area_start_row + 1 + a_idx
        ws_ov.cell(row=curr_row, column=1, value=row['Division']).alignment = ALIGN_CENTER
        ws_ov.cell(row=curr_row, column=2, value=row['Area']).alignment = ALIGN_CENTER
        
        ws_ov.cell(row=curr_row, column=3, value=f"=SUMIFS(ClubDetails!G:G, ClubDetails!C:C, A{curr_row}, ClubDetails!D:D, B{curr_row})")
        ws_ov.cell(row=curr_row, column=4, value=f"=SUMIFS(ClubDetails!H:H, ClubDetails!C:C, A{curr_row}, ClubDetails!D:D, B{curr_row})")
        ws_ov.cell(row=curr_row, column=5, value=f"=SUMIFS(ClubDetails!K:K, ClubDetails!C:C, A{curr_row}, ClubDetails!D:D, B{curr_row})")
        ws_ov.cell(row=curr_row, column=6, value=f"=IF(C{curr_row}=0, 0, E{curr_row}/C{curr_row})")
        ws_ov.cell(row=curr_row, column=7, value=f"=SUMIFS(ClubDetails!L:L, ClubDetails!C:C, A{curr_row}, ClubDetails!D:D, B{curr_row})")
        ws_ov.cell(row=curr_row, column=8, value=f"=IF(C{curr_row}=0, 0, G{curr_row}/C{curr_row})")
        ws_ov.cell(row=curr_row, column=9, value=f"=SUMIFS(ClubDetails!J:J, ClubDetails!C:C, A{curr_row}, ClubDetails!D:D, B{curr_row})")
        ws_ov.cell(row=curr_row, column=10, value=f"=SUMIFS(ClubDetails!M:M, ClubDetails!C:C, A{curr_row}, ClubDetails!D:D, B{curr_row})")
        
        fill = ZEBRA_FILL if curr_row % 2 == 0 else WHITE_FILL
        for c_idx in range(1, 11):
            cell = ws_ov.cell(row=curr_row, column=c_idx)
            cell.font = FONT_BODY
            cell.border = THIN_BORDER
            cell.fill = fill
            if c_idx > 2:
                cell.alignment = ALIGN_RIGHT
                if c_idx in [6, 8]:
                    cell.number_format = "0.0%"
                else:
                    cell.number_format = "#,##0"
                    
    area_tot_row = area_start_row + len(area_combinations) + 1
    ws_ov.cell(row=area_tot_row, column=1, value="Area Total").font = FONT_BODY_BOLD
    ws_ov.cell(row=area_tot_row, column=1).alignment = ALIGN_LEFT
    ws_ov.cell(row=area_tot_row, column=1).border = THIN_BORDER
    ws_ov.cell(row=area_tot_row, column=1).fill = SILVER_FILL
    ws_ov.cell(row=area_tot_row, column=2, value="").border = THIN_BORDER
    ws_ov.cell(row=area_tot_row, column=2).fill = SILVER_FILL
    
    ws_ov.cell(row=area_tot_row, column=3, value=f"=SUM(C{area_start_row+1}:C{area_tot_row-1})")
    ws_ov.cell(row=area_tot_row, column=4, value=f"=SUM(D{area_start_row+1}:D{area_tot_row-1})")
    ws_ov.cell(row=area_tot_row, column=5, value=f"=SUM(E{area_start_row+1}:E{area_tot_row-1})")
    ws_ov.cell(row=area_tot_row, column=6, value=f"=IF(C{area_tot_row}=0, 0, E{area_tot_row}/C{area_tot_row})")
    ws_ov.cell(row=area_tot_row, column=7, value=f"=SUM(G{area_start_row+1}:G{area_tot_row-1})")
    ws_ov.cell(row=area_tot_row, column=8, value=f"=IF(C{area_tot_row}=0, 0, G{area_tot_row}/C{area_tot_row})")
    ws_ov.cell(row=area_tot_row, column=9, value=f"=SUM(I{area_start_row+1}:I{area_tot_row-1})")
    ws_ov.cell(row=area_tot_row, column=10, value=f"=SUM(J{area_start_row+1}:J{area_tot_row-1})")
    
    for c_idx in range(3, 11):
        cell = ws_ov.cell(row=area_tot_row, column=c_idx)
        cell.font = FONT_BODY_BOLD
        cell.border = THIN_BORDER
        cell.fill = SILVER_FILL
        cell.alignment = ALIGN_RIGHT
        if c_idx in [6, 8]:
            cell.number_format = "0.0%"
        else:
            cell.number_format = "#,##0"

    ws_ov.column_dimensions['A'].width = 15
    autofit_columns(ws_ov)

    # 4. District Dashboard Sheet (Index 0)
    ws_dist = wb.create_sheet(title="District", index=0)
    ws_dist.views.sheetView[0].showGridLines = True
    
    ws_dist.cell(row=1, column=1, value="District 227 Performance Dashboard").font = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
    ws_dist.cell(row=2, column=1, value="Membership, Payments & Club Recognition Goals").font = FONT_SUBTITLE
    
    ws_dist.cell(row=4, column=1, value="Payments Goals").font = FONT_BODY_BOLD
    ws_dist.cell(row=4, column=1).fill = LIGHT_NAVY_FILL
    ws_dist.cell(row=4, column=1).font = FONT_HEADER_WHITE
    ws_dist.merge_cells("A4:B4")
    
    ws_dist.cell(row=4, column=4, value="Clubs Goals").font = FONT_BODY_BOLD
    ws_dist.cell(row=4, column=4).fill = LIGHT_NAVY_FILL
    ws_dist.cell(row=4, column=4).font = FONT_HEADER_WHITE
    ws_dist.merge_cells("D4:E4")
    
    ws_dist.cell(row=4, column=7, value="Distinguished Clubs Goals").font = FONT_BODY_BOLD
    ws_dist.cell(row=4, column=7).fill = LIGHT_NAVY_FILL
    ws_dist.cell(row=4, column=7).font = FONT_HEADER_WHITE
    ws_dist.merge_cells("G4:H4")
    
    goals_data = [
        ("Base Payments", 8634, "Base Clubs", 174, "Distinguished Clubs Prereq", 0),
        ("Distinguished", "=ROUND(B5*1.01, 0)", "Distinguished", "=E5", "Distinguished", "=ROUND(0.4*E5, 0)"),
        ("Select Distinguished", "=ROUND(B5*1.03, 0)", "Select Distinguished", 184, "Select Distinguished", 89),
        ("President's Distinguished", "=ROUND(B5*1.05, 0)", "President's Distinguished", 194, "President's Distinguished", 100),
        ("Smedley Distinguished", "=ROUND(B5*1.08, 0)", "Smedley Distinguished", 197, "Smedley Distinguished", 108),
    ]
    
    for r_idx, row in enumerate(goals_data, 5):
        ws_dist.cell(row=r_idx, column=1, value=row[0]).font = FONT_BODY
        ws_dist.cell(row=r_idx, column=2, value=row[1]).font = FONT_BODY_BOLD
        ws_dist.cell(row=r_idx, column=2).alignment = ALIGN_RIGHT
        
        ws_dist.cell(row=r_idx, column=4, value=row[2]).font = FONT_BODY
        ws_dist.cell(row=r_idx, column=5, value=row[3]).font = FONT_BODY_BOLD
        ws_dist.cell(row=r_idx, column=5).alignment = ALIGN_RIGHT
        
        ws_dist.cell(row=r_idx, column=7, value=row[4]).font = FONT_BODY
        ws_dist.cell(row=r_idx, column=8, value=row[5]).font = FONT_BODY_BOLD
        ws_dist.cell(row=r_idx, column=8).alignment = ALIGN_RIGHT
        
        ws_dist.cell(row=r_idx, column=1).border = THIN_BORDER
        ws_dist.cell(row=r_idx, column=2).border = THIN_BORDER
        ws_dist.cell(row=r_idx, column=4).border = THIN_BORDER
        ws_dist.cell(row=r_idx, column=5).border = THIN_BORDER
        ws_dist.cell(row=r_idx, column=7).border = THIN_BORDER
        ws_dist.cell(row=r_idx, column=8).border = THIN_BORDER
        
        if not str(row[1]).startswith("="): ws_dist.cell(row=r_idx, column=2).number_format = "#,##0"
        if not str(row[3]).startswith("="): ws_dist.cell(row=r_idx, column=5).number_format = "#,##0"
        if not str(row[5]).startswith("="): ws_dist.cell(row=r_idx, column=8).number_format = "#,##0"
        
    ws_dist["B6"].number_format = "#,##0"
    ws_dist["B7"].number_format = "#,##0"
    ws_dist["B8"].number_format = "#,##0"
    ws_dist["B9"].number_format = "#,##0"
    ws_dist["E6"].number_format = "#,##0"
    ws_dist["H6"].number_format = "#,##0"
    
    tot_row = len(divisions) + 4
    ws_dist.cell(row=11, column=1, value="Payments To Date").font = FONT_BODY_BOLD
    ws_dist.cell(row=11, column=2, value=f"=Overall!L{tot_row}").font = FONT_BODY_BOLD
    ws_dist.cell(row=11, column=2).alignment = ALIGN_RIGHT
    ws_dist.cell(row=11, column=2).number_format = "#,##0"
    ws_dist.cell(row=11, column=2).fill = SILVER_FILL
    ws_dist.cell(row=11, column=1).border = THIN_BORDER
    ws_dist.cell(row=11, column=2).border = THIN_BORDER
    
    ws_dist.cell(row=11, column=4, value="Active Clubs").font = FONT_BODY_BOLD
    ws_dist.cell(row=11, column=5, value=f"=Overall!C{tot_row}").font = FONT_BODY_BOLD
    ws_dist.cell(row=11, column=5).alignment = ALIGN_RIGHT
    ws_dist.cell(row=11, column=5).number_format = "#,##0"
    ws_dist.cell(row=11, column=5).fill = SILVER_FILL
    ws_dist.cell(row=11, column=4).border = THIN_BORDER
    ws_dist.cell(row=11, column=5).border = THIN_BORDER
    
    ws_dist.cell(row=11, column=7, value="Distinguished To Date").font = FONT_BODY_BOLD
    ws_dist.cell(row=11, column=8, value=f"=Overall!D{tot_row}").font = FONT_BODY_BOLD
    ws_dist.cell(row=11, column=8).alignment = ALIGN_RIGHT
    ws_dist.cell(row=11, column=8).number_format = "#,##0"
    ws_dist.cell(row=11, column=8).fill = SILVER_FILL
    ws_dist.cell(row=11, column=7).border = THIN_BORDER
    ws_dist.cell(row=11, column=8).border = THIN_BORDER

    pending_goals = [
        ("Pending Distinguished", "=IF(B6-B11<0,0,B6-B11)", "Pending Distinguished", "=IF(E6-E11<0,0,E6-E11)", "Pending Distinguished", "=IF(H6-H11<0,0,H6-H11)"),
        ("Pending Select Distinguished", "=IF(B7-B11<0,0,B7-B11)", "Pending Select Distinguished", "=IF(E7-E11<0,0,E7-E11)", "Pending Select Distinguished", "=IF(H7-H11<0,0,H7-H11)"),
        ("Pending President's Distinguished", "=IF(B8-B11<0,0,B8-B11)", "Pending President's Distinguished", "=IF(E8-E11<0,0,E8-E11)", "Pending President's Distinguished", "=IF(H8-H11<0,0,H8-H11)"),
        ("Pending Smedley Distinguished", "=IF(B9-B11<0,0,B9-B11)", "Pending Smedley Distinguished", "=IF(E9-E11<0,0,E9-E11)", "Pending Smedley Distinguished", "=IF(H9-H11<0,0,H9-H11)"),
    ]
    
    for p_idx, row in enumerate(pending_goals, 13):
        ws_dist.cell(row=p_idx, column=1, value=row[0]).font = FONT_BODY
        ws_dist.cell(row=p_idx, column=2, value=row[1]).font = FONT_BODY
        ws_dist.cell(row=p_idx, column=2).alignment = ALIGN_RIGHT
        ws_dist.cell(row=p_idx, column=2).number_format = "#,##0"
        
        ws_dist.cell(row=p_idx, column=4, value=row[2]).font = FONT_BODY
        ws_dist.cell(row=p_idx, column=5, value=row[3]).font = FONT_BODY
        ws_dist.cell(row=p_idx, column=5).alignment = ALIGN_RIGHT
        ws_dist.cell(row=p_idx, column=5).number_format = "#,##0"
        
        ws_dist.cell(row=p_idx, column=7, value=row[4]).font = FONT_BODY
        ws_dist.cell(row=p_idx, column=8, value=row[5]).font = FONT_BODY
        ws_dist.cell(row=p_idx, column=8).alignment = ALIGN_RIGHT
        ws_dist.cell(row=p_idx, column=8).number_format = "#,##0"
        
        ws_dist.cell(row=p_idx, column=1).border = THIN_BORDER
        ws_dist.cell(row=p_idx, column=2).border = THIN_BORDER
        ws_dist.cell(row=p_idx, column=4).border = THIN_BORDER
        ws_dist.cell(row=p_idx, column=5).border = THIN_BORDER
        ws_dist.cell(row=p_idx, column=7).border = THIN_BORDER
        ws_dist.cell(row=p_idx, column=8).border = THIN_BORDER

    ws_dist.cell(row=18, column=1, value="Payments % Base").font = FONT_BODY_BOLD
    ws_dist.cell(row=18, column=2, value="=B11/B5").font = FONT_BODY_BOLD
    ws_dist.cell(row=18, column=2).alignment = ALIGN_RIGHT
    ws_dist.cell(row=18, column=2).number_format = "0.0%"
    ws_dist.cell(row=18, column=1).border = THIN_BORDER
    ws_dist.cell(row=18, column=2).border = THIN_BORDER
    
    ws_dist.cell(row=19, column=1, value="Payments % SDD Goal").font = FONT_BODY_BOLD
    ws_dist.cell(row=19, column=2, value="=B11/B9").font = FONT_BODY_BOLD
    ws_dist.cell(row=19, column=2).alignment = ALIGN_RIGHT
    ws_dist.cell(row=19, column=2).number_format = "0.0%"
    ws_dist.cell(row=19, column=1).border = THIN_BORDER
    ws_dist.cell(row=19, column=2).border = THIN_BORDER
    
    ws_dist.cell(row=18, column=4, value="20 + Clubs").font = FONT_BODY_BOLD
    ws_dist.cell(row=18, column=5, value=f"=Overall!M{tot_row}").font = FONT_BODY_BOLD
    ws_dist.cell(row=18, column=5).alignment = ALIGN_RIGHT
    ws_dist.cell(row=18, column=5).number_format = "#,##0"
    ws_dist.cell(row=18, column=4).border = THIN_BORDER
    ws_dist.cell(row=18, column=5).border = THIN_BORDER
    
    ws_dist.cell(row=19, column=4, value="20 + Clubs % Active Clubs").font = FONT_BODY_BOLD
    ws_dist.cell(row=19, column=5, value="=E18/E11").font = FONT_BODY_BOLD
    ws_dist.cell(row=19, column=5).alignment = ALIGN_RIGHT
    ws_dist.cell(row=19, column=5).number_format = "0.0%"
    ws_dist.cell(row=19, column=4).border = THIN_BORDER
    ws_dist.cell(row=19, column=5).border = THIN_BORDER
    
    ws_dist.cell(row=18, column=7, value="Distinguished % Base").font = FONT_BODY_BOLD
    ws_dist.cell(row=18, column=8, value="=H11/E5").font = FONT_BODY_BOLD
    ws_dist.cell(row=18, column=8).alignment = ALIGN_RIGHT
    ws_dist.cell(row=18, column=8).number_format = "0.0%"
    ws_dist.cell(row=18, column=7).border = THIN_BORDER
    ws_dist.cell(row=18, column=8).border = THIN_BORDER

    ws_dist.cell(row=21, column=1, value="Membership Campaigns & Deadlines").font = Font(name="Segoe UI", size=12, bold=True, color="1B365D")
    
    campaign_headers = ["Campaign Name", "Timeline", "Status / Time Left"]
    for c_idx, h in enumerate(campaign_headers, 1):
        cell = ws_dist.cell(row=23, column=c_idx, value=h)
        cell.font = FONT_HEADER_WHITE
        cell.fill = LIGHT_NAVY_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        
    campaigns = [
        ("Smedley Award", "Aug 1 - Sep 30", '=IF(TODAY()>DATE(YEAR(TODAY()),9,30), "Ended", DATE(YEAR(TODAY()),9,30)-TODAY() & " days left")'),
        ("Talk Up Toastmasters", "Feb 1 - Mar 31", '=IF(TODAY()>DATE(YEAR(TODAY())+IF(MONTH(TODAY())>3,1,0),3,31), "Ended", DATE(YEAR(TODAY())+IF(MONTH(TODAY())>3,1,0),3,31)-TODAY() & " days left")'),
        ("Beat the Clock", "May 1 - Jun 30", '=IF(TODAY()>DATE(YEAR(TODAY())+IF(MONTH(TODAY())>6,1,0),6,30), "Ended", DATE(YEAR(TODAY())+IF(MONTH(TODAY())>6,1,0),6,30)-TODAY() & " days left")'),
    ]
    
    for r_offset, camp in enumerate(campaigns, 24):
        ws_dist.cell(row=r_offset, column=1, value=camp[0]).font = FONT_BODY
        ws_dist.cell(row=r_offset, column=2, value=camp[1]).font = FONT_BODY
        ws_dist.cell(row=r_offset, column=3, value=camp[2]).font = FONT_BODY_BOLD
        
        ws_dist.cell(row=r_offset, column=1).border = THIN_BORDER
        ws_dist.cell(row=r_offset, column=2).border = THIN_BORDER
        ws_dist.cell(row=r_offset, column=3).border = THIN_BORDER
        ws_dist.cell(row=r_offset, column=3).alignment = ALIGN_CENTER
        ws_dist.cell(row=r_offset, column=2).alignment = ALIGN_CENTER

    ws_dist.column_dimensions['A'].width = 30
    ws_dist.column_dimensions['B'].width = 15
    ws_dist.column_dimensions['C'].width = 25
    ws_dist.column_dimensions['D'].width = 30
    ws_dist.column_dimensions['E'].width = 15
    ws_dist.column_dimensions['F'].width = 5
    ws_dist.column_dimensions['G'].width = 30
    ws_dist.column_dimensions['H'].width = 15
    
    wb.save(output_excel_path)
    logging.info(f"✅ Master Excel Workbook generated with all sheets: {output_excel_path}")

def process_downloaded_data(data_dir, workspace_dir, timestamp):
    logging.info("📊 Processing downloaded performance data into Master Excel Workbook...")
    final_df, raw_files = load_and_clean_data(data_dir, timestamp)

    if final_df is None:
        return

    output_excel_path = os.path.join(workspace_dir, "District 227 - Mastersheet.xlsx")
    generate_excel_mastersheet(final_df, output_excel_path)

    output_json_path = os.path.join(workspace_dir, "District 227 - Mastersheet.json")
    final_df.to_json(output_json_path, orient="records", indent=2)

    # Save copies inside dashboard directory for direct automatic serving
    dashboard_dir = os.path.join(workspace_dir, "dashboard")
    if os.path.exists(dashboard_dir):
        dashboard_excel_path = os.path.join(dashboard_dir, "District 227 - Mastersheet.xlsx")
        shutil.copy2(output_excel_path, dashboard_excel_path)
        
        dashboard_json_path = os.path.join(dashboard_dir, "District 227 - Mastersheet.json")
        final_df.to_json(dashboard_json_path, orient="records", indent=2)
        logging.info(f"✅ Copied Mastersheet Excel & JSON to dashboard folder.")
        
        # Mirror static web assets to root for GitHub Pages root hosting
        for fname in ["index.html", "style.css", "app.js", "images.png", "toastmasters logo.jpg"]:
            src = os.path.join(dashboard_dir, fname)
            dst = os.path.join(workspace_dir, fname)
            if os.path.exists(src):
                shutil.copy2(src, dst)

    # Clean raw temp CSV files
    for f in raw_files:
        try:
            os.remove(f)
            logging.info(f"🗑️ Cleaned temp file: {f}")
        except Exception as e:
            logging.warning(f"Could not remove {f}: {e}")

def run_toastmasters_pipeline():
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    logging.info("🚀 Starting Toastmasters Download & Excel Mastersheet Pipeline...")
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        
        for target_name, target_url in TARGET_PAGES.items():
            try:
                download_data_from_url(page, target_name, target_url, timestamp)
            except Exception as e:
                logging.error(f"❌ Failed downloading {target_name}: {str(e)}")
                
        browser.close()
        
    data_dir = os.path.join(os.getcwd(), "data")
    process_downloaded_data(data_dir, os.getcwd(), timestamp)
    logging.info("🎉 Pipeline execution complete!")

if __name__ == "__main__":
    setup_logging()
    run_toastmasters_pipeline()
